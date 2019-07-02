#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl, argparse, time
from colorama import init, Fore, Back, Style
from googletrans import Translator
from progressbar import ProgressBar, Counter, Timer

const01 = "Traduci la prima colonna del file excel da 'it' a xx. Verrà utilizzato Google Translate"
const04 = 'Caricamento del file'
const05 = 'Scrittura del file'
const02 = '[FILE] #Inserisci il percorso del file o il nome, se si trova nella stessa directory dello script '
const03 = '[XX]   #Inserisci il codice della lingua. '
texte01 = Fore.RED + 'ERRORE:' + Fore.RESET + 'Problema nella lettura del file'
texte02 = Fore.RED + 'ERRORE:' + Fore.RESET + 'Inserire il percorso del file'
texte03 = Fore.RED + 'ERRORE:' + Fore.RESET + 'Problema nella traduzione. Verifca che la lingua di destinazioe sia corretta. Oppure hai fatto troppi tentativi. Attendere'

trans = Translator(service_urls=['translate.google.com'])

# Per attivazione colori su Windows
init(convert=True)

def translate_column_value(parola_from, lingua):
        # Lingua di partenza it, lingua di destinazione quella messa a parametro
        translation = trans.translate(parola_from, src='it', dest=lingua)
        parola_to = translation.text
        return parola_to

def scrivi_file_excel(excel, lingua):
    # Prendo il primo foglio attivo
    pysheet = excel.active

    # Prendo tutta la prima colonna
    column = list(pysheet.columns)[0]
    rows = pysheet.max_row

    idx = 0
    # Imposto la barra di esecuzione
    widgets = ['Processed: ', Counter(), ' lines (', Timer(), ')']
    pbar = ProgressBar(widgets=widgets)
    for i in pbar((i for i in range(rows))):
        # Per ogni colla vado ad effettuare la traduzione
        time.sleep(2)
        cell = column[i]
        idx = idx + 1
        try:
            if idx == 100:
                time.sleep(10)
            # Lancio la funzione per tradurre il valore della cella
            parola_trad = translate_column_value(cell.value, lingua)
            cella1 = pysheet.cell(row=idx, column=2)
            # Scrivo il valore della parola tradotta nella nuova cella
            cella1.value = parola_trad
        except:
            # Se ho un errore nella traduzione
            print(texte03)
            break
    return

def main_menu(percorsofile, lingua):
    try:
        print(const04)
        excel = openpyxl.load_workbook(percorsofile)
        print(const05)
        scrivi_file_excel(excel, lingua)
        excel.save(percorsofile)
    except:
        # Errori nella lettura del file
        if not percorsofile:
            print(texte02)
        else:
            print(texte01)


if __name__ == "__main__":

    print(const01)
    f = input(const02)
    l = input(const03)

    if not (args.l):
        args.l = 'en'

    main_menu(f, l)

